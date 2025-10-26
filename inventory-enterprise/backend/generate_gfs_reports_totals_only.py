#!/usr/bin/env python3
"""
GFS Monthly Accounting Report Generator - Invoice Totals Only
Generates reports by fiscal period showing invoice totals without category breakdown
Due to line items data quality issues, this version shows invoice totals only.
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# ============================================================================
# CONFIGURATION
# ============================================================================

DATABASE_PATH = "data/enterprise_inventory.db"
OUTPUT_DIR = "/Users/davidmikulis/Desktop/GFS_Fiscal_Reports"
PDF_BASE_DIR = "/Users/davidmikulis/Desktop/Invoices"

# Simplified 10-column schema (no category breakdown)
COLUMNS = [
    "Fiscal Period",
    "Week Ending",
    "Vendor",
    "Date",
    "Invoice #",
    "Total Invoice Amount",
    "63107000 GST",
    "63107100 QST",
    "📎 Document Link",
    "Notes"
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def create_pdf_hyperlink(invoice_num):
    """Create Excel HYPERLINK formula for PDF document"""
    if not invoice_num:
        return "NO PDF FOUND"

    invoice_clean = str(invoice_num).replace(' ', '').replace('#', '').split('.')[0]

    # Use API endpoint for PDF viewing
    url = f"http://localhost:8083/api/owner/docs/view/{invoice_clean}"
    formula = f'=HYPERLINK("{url}","📄 {invoice_clean}")'

    return formula

def safe_float(val):
    """Convert value to float, return 0.0 if invalid"""
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def get_fiscal_periods(conn):
    """Query database to find all fiscal periods with invoices"""
    query = """
    SELECT DISTINCT
        fp.fiscal_year,
        fp.period,
        fp.fiscal_year_id,
        fp.period_start_date,
        fp.period_end_date,
        fp.notes as period_name
    FROM fiscal_periods fp
    JOIN documents d ON d.fiscal_year_id = fp.fiscal_year_id
        AND d.fiscal_period_id = 'FY' || (fp.fiscal_year % 100) || '-P' || printf('%02d', fp.period)
    WHERE d.mime_type = 'application/pdf'
      AND d.deleted_at IS NULL
    ORDER BY fp.fiscal_year, fp.period
    """

    df = pd.read_sql_query(query, conn)
    periods = []

    for _, row in df.iterrows():
        periods.append({
            'fiscal_year': int(row['fiscal_year']),
            'period': int(row['period']),
            'fiscal_year_id': row['fiscal_year_id'],
            'period_id': f"FY{int(row['fiscal_year']) % 100}-P{int(row['period']):02d}",
            'period_name': row['period_name'],
            'start_date': row['period_start_date'],
            'end_date': row['period_end_date']
        })

    return periods

# ============================================================================
# REPORT GENERATION
# ============================================================================

def generate_fiscal_period_report(period_info):
    """Generate report for a specific fiscal period"""

    fiscal_year = period_info['fiscal_year']
    period = period_info['period']
    period_id = period_info['period_id']
    period_name = period_info['period_name']
    start_date = period_info['start_date']
    end_date = period_info['end_date']

    sheet_name = f"{period_id}_{period_name.replace(' ', '_')}"

    print(f"\n{'='*80}")
    print(f"Generating Report: {period_id} - {period_name}")
    print(f"{'='*80}")
    print(f"Date Range: {start_date} to {end_date}")

    # Connect to database
    conn = sqlite3.connect(DATABASE_PATH)

    # Query invoices for this fiscal period
    query_invoices = """
    SELECT
        d.id as invoice_id,
        d.invoice_number,
        d.vendor as supplier,
        d.invoice_date,
        d.invoice_amount as total_amount,
        d.fiscal_year_id,
        d.fiscal_period_id
    FROM documents d
    WHERE d.fiscal_period_id = ?
      AND d.mime_type = 'application/pdf'
      AND d.deleted_at IS NULL
    ORDER BY d.invoice_date, d.invoice_number
    """

    df_invoices = pd.read_sql_query(query_invoices, conn, params=(period_id,))
    print(f"  ✓ Found {len(df_invoices)} invoices")

    if len(df_invoices) == 0:
        print(f"  ⚠️  No invoices for {period_id} - {period_name}")
        conn.close()
        return None

    conn.close()

    # Process invoices
    output_rows = []

    for idx, invoice_row in df_invoices.iterrows():
        invoice_id = invoice_row['invoice_id']
        invoice_number = invoice_row['invoice_number']
        supplier = invoice_row['supplier'] or 'GFS'
        invoice_date = invoice_row['invoice_date']
        total_amount = safe_float(invoice_row['total_amount'])

        # Create row with just invoice totals
        row = {
            'Fiscal Period': period_id,
            'Week Ending': end_date,  # Use fiscal period end date
            'Vendor': supplier,
            'Date': invoice_date,
            'Invoice #': invoice_number,
            'Total Invoice Amount': round(total_amount, 2),
            '63107000 GST': 0.0,  # Not available from line items
            '63107100 QST': 0.0,  # Not available from line items
            '📎 Document Link': create_pdf_hyperlink(invoice_number),
            'Notes': 'Category breakdown not available - line items data requires correction'
        }

        output_rows.append(row)

    print(f"  ✓ Processed {len(output_rows)} invoice rows")

    # Create dataframe
    df_output = pd.DataFrame(output_rows, columns=COLUMNS)

    # Round numeric columns
    numeric_cols = ['Total Invoice Amount', '63107000 GST', '63107100 QST']
    for col in numeric_cols:
        df_output[col] = df_output[col].apply(lambda x: round(float(x), 2) if pd.notna(x) else 0.0)

    # Export to Excel
    output_filename = os.path.join(OUTPUT_DIR, f"GFS_Totals_{period_id}_{period_name.replace(' ', '_')}.xlsx")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        # Write headers
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_idx, row_data in enumerate(df_output.to_dict('records'), start=2):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)

                if col_name == '📎 Document Link':
                    cell.value = value
                elif col_name in numeric_cols:
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.value = value

        # Add grand total row
        total_row_idx = len(df_output) + 2
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws.cell(row=total_row_idx, column=1, value="GRAND TOTAL").font = total_font
        ws.cell(row=total_row_idx, column=1).fill = total_fill

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            if col_name in numeric_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                formula = f'=SUM({col_letter}2:{col_letter}{total_row_idx-1})'
                cell.value = formula
                cell.font = total_font
                cell.fill = total_fill
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

        # Adjust column widths
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name == '📎 Document Link':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
            elif col_name == 'Notes':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 50
            elif col_name in ['Vendor', 'Invoice #', 'Fiscal Period']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            elif col_name in numeric_cols:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

        # Save workbook
        wb.save(output_filename)
        print(f"  ✓ Saved: {output_filename}")

        # Return summary
        total_amount = df_output['Total Invoice Amount'].sum()
        return {
            'fiscal_year': fiscal_year,
            'period': period,
            'period_id': period_id,
            'period_name': period_name,
            'invoice_count': len(df_output),
            'total_amount': total_amount,
            'filename': output_filename
        }

    except Exception as e:
        print(f"  ✗ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

# ============================================================================
# MAIN FUNCTION
# ============================================================================

def generate_all_fiscal_reports():
    """Generate reports for all fiscal periods with invoices"""

    print("=" * 80)
    print("GFS FISCAL PERIOD ACCOUNTING REPORT GENERATOR")
    print("Invoice Totals Only (Category Breakdown Not Available)")
    print("=" * 80)
    print()

    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Output Directory: {OUTPUT_DIR}")

    # Connect to database
    print("\nConnecting to database...")
    conn = sqlite3.connect(DATABASE_PATH)

    # Get all fiscal periods with invoices
    print("Scanning database for fiscal periods...")
    fiscal_periods = get_fiscal_periods(conn)
    conn.close()

    if len(fiscal_periods) == 0:
        print("✗ No invoices found in database")
        return

    print(f"✓ Found invoices in {len(fiscal_periods)} fiscal periods:")
    for period in fiscal_periods:
        print(f"  • {period['period_id']}: {period['period_name']} ({period['start_date']} to {period['end_date']})")

    # Generate reports
    print("\n" + "=" * 80)
    print("GENERATING FISCAL PERIOD REPORTS")
    print("=" * 80)

    summaries = []
    for period in fiscal_periods:
        summary = generate_fiscal_period_report(period)
        if summary:
            summaries.append(summary)

    # Final summary
    print("\n" + "=" * 80)
    print("GENERATION COMPLETE")
    print("=" * 80)
    print()

    if summaries:
        total_invoices = sum(s['invoice_count'] for s in summaries)
        total_amount = sum(s['total_amount'] for s in summaries)

        print(f"✅ Generated {len(summaries)} fiscal period reports")
        print(f"📊 Total Invoices: {total_invoices}")
        print(f"💰 Total Amount: ${total_amount:,.2f}")
        print()

        print("FISCAL PERIOD BREAKDOWN:")
        print("-" * 80)
        for s in summaries:
            print(f"  {s['period_id']} {s['period_name']:20s}  |  {s['invoice_count']:3d} invoices  |  ${s['total_amount']:12,.2f}")
        print("-" * 80)
        print()

        print(f"📁 All reports saved to: {OUTPUT_DIR}")
        print()
        print("NOTES:")
        print("  • Reports show invoice totals only")
        print("  • Category breakdown not available due to line items data quality issues")
        print("  • Invoice totals are accurate for accounting purposes")
        print("  • Click PDF links to view source documents")
    else:
        print("✗ No reports generated")

if __name__ == "__main__":
    import sys
    try:
        generate_all_fiscal_reports()
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

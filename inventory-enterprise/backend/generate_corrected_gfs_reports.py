#!/usr/bin/env python3
"""
QUICK FIX: Generate Corrected GFS Reports
Uses line_total directly (not unit_price * quantity) and applies category mapping
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

DB_PATH = "data/enterprise_inventory.db"
OUTPUT_DIR = os.path.expanduser("~/Desktop/GFS_Fiscal_Reports_CORRECTED")

# Category mapping rules (based on GFS category names and keywords)
CATEGORY_MAP = {
    # GFS category name mappings
    'Produce': '60110070 PROD',
    'Meat': '60110060 MEAT',
    'Poultry': '60110060 MEAT',
    'Seafood': '60110060 MEAT',
    'Dairy': '60110030 MILK',
    'Frozen': '60110040 GROC + MISC',
    'Grocery': '60110040 GROC + MISC',
    'Beverage': '60110020 BEV + ECO',
    'Disposables': '60260010 PAPER',
    'Chemical': '60220001 CLEAN',
    'Tabletop': '60665001 Small Equip',
    'Fuel Charge': '62421100 FREIGHT',
}

# Keyword-based mapping (fallback)
KEYWORD_MAP = {
    'BEEF': '60110060 MEAT',
    'PORK': '60110060 MEAT',
    'CHICKEN': '60110060 MEAT',
    'MEAT': '60110060 MEAT',
    'FISH': '60110060 MEAT',
    'SEAFOOD': '60110060 MEAT',
    'SALMON': '60110060 MEAT',
    'HADDOCK': '60110060 MEAT',
    'BACON': '60110060 MEAT',
    'SAUSAGE': '60110060 MEAT',
    'HAM': '60110060 MEAT',
    'BOLOGNA': '60110060 MEAT',
    'PASTRAMI': '60110060 MEAT',
    'MEATBALL': '60110060 MEAT',

    'APPLE': '60110070 PROD',
    'BANANA': '60110070 PROD',
    'ORANGE': '60110070 PROD',
    'PRODUCE': '60110070 PROD',
    'LETTUCE': '60110070 PROD',
    'TOMATO': '60110070 PROD',
    'CARROT': '60110070 PROD',
    'CELERY': '60110070 PROD',
    'BROCCOLI': '60110070 PROD',
    'CAULIFLOWER': '60110070 PROD',
    'CABBAGE': '60110070 PROD',
    'FRUIT': '60110070 PROD',
    'BERRY': '60110070 PROD',
    'BEET': '60110070 PROD',

    'MILK': '60110030 MILK',
    'CHEESE': '60110030 MILK',
    'YOGURT': '60110030 MILK',
    'CREAM': '60110030 MILK',
    'BUTTER': '60110030 MILK',
    'DAIRY': '60110030 MILK',
    'MOZZ': '60110030 MILK',
    'EGG': '60110030 MILK',

    'BREAD': '60110010 BAKE',
    'ROLL': '60110010 BAKE',
    'BUN': '60110010 BAKE',
    'BAGEL': '60110010 BAKE',
    'MUFFIN': '60110010 BAKE',

    'JUICE': '60110020 BEV + ECO',
    'COFFEE': '60110020 BEV + ECO',
    'TEA': '60110020 BEV + ECO',

    'PAPER': '60260010 PAPER',
    'NAPKIN': '60260010 PAPER',
    'CUP': '60260010 PAPER',
    'PLATE': '60260010 PAPER',
    'PPR': '60260010 PAPER',

    'CLEAN': '60220001 CLEAN',
    'SOAP': '60220001 CLEAN',
    'SANITIZER': '60220001 CLEAN',
    'CHEMICAL': '60220001 CLEAN',

    'FREIGHT': '62421100 FREIGHT',
    'FUEL': '62421100 FREIGHT',
}

def map_category(description, category_name=None):
    """Map item to category code"""

    # Try GFS category name first
    if category_name and category_name in CATEGORY_MAP:
        return CATEGORY_MAP[category_name]

    # Try keyword matching on description
    if description:
        desc_upper = str(description).upper()
        for keyword, cat_code in KEYWORD_MAP.items():
            if keyword in desc_upper:
                return cat_code

    # Default to OTHER
    return 'Other Costs'

def generate_report(period_id):
    """Generate corrected report for a fiscal period"""

    print("=" * 80)
    print(f"CORRECTED GFS REPORT - {period_id}")
    print("=" * 80)
    print()

    conn = sqlite3.connect(DB_PATH)

    # Get invoices for this period
    df_invoices = pd.read_sql_query("""
        SELECT
            d.invoice_number,
            d.invoice_date,
            d.vendor,
            d.invoice_amount as total_amount
        FROM documents d
        WHERE d.fiscal_period_id = ?
          AND d.mime_type = 'application/pdf'
          AND d.deleted_at IS NULL
        ORDER BY d.invoice_date, d.invoice_number
    """, conn, params=(period_id,))

    if len(df_invoices) == 0:
        print(f"❌ No invoices found for period {period_id}")
        conn.close()
        return None

    print(f"✓ Found {len(df_invoices)} invoices")

    # Get line items - USE line_total field directly (don't calculate!)
    df_lines = pd.read_sql_query("""
        SELECT
            ili.invoice_number,
            ili.description,
            ili.category,
            ili.line_total,
            ili.quantity,
            ili.unit_price
        FROM invoice_line_items ili
        WHERE ili.invoice_number IN ({})
    """.format(','.join(['?'] * len(df_invoices))), conn, params=df_invoices['invoice_number'].tolist())

    print(f"✓ Found {len(df_lines)} line items")

    # Map categories
    df_lines['category_code'] = df_lines.apply(
        lambda row: map_category(row['description'], row['category']),
        axis=1
    )

    # Create category summary
    category_totals = {}
    for _, row in df_lines.iterrows():
        cat = row['category_code']
        amount = float(row['line_total']) if pd.notna(row['line_total']) else 0.0
        category_totals[cat] = category_totals.get(cat, 0.0) + amount

    print()
    print("CATEGORY BREAKDOWN:")
    print("-" * 80)
    for cat, total in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
        print(f"  {cat:30s}  ${total:12,.2f}")

    print()
    print(f"Total from line items: ${sum(category_totals.values()):,.2f}")
    print(f"Total from invoices:   ${df_invoices['total_amount'].sum():,.2f}")

    variance = abs(sum(category_totals.values()) - df_invoices['total_amount'].sum())
    print(f"Variance:              ${variance:,.2f}")

    if variance > 100:
        print()
        print(f"⚠️  WARNING: Large variance detected (${variance:,.2f})")
        print("   This may indicate missing line items or taxes included in totals")

    # Pivot data for report
    invoice_rows = []
    for _, inv in df_invoices.iterrows():
        invoice_num = inv['invoice_number']
        invoice_lines = df_lines[df_lines['invoice_number'] == invoice_num]

        row = {
            'Invoice #': invoice_num,
            'Date': inv['invoice_date'],
            'Vendor': inv['vendor'] or 'GFS',
            'Total Invoice Amount': inv['total_amount']
        }

        # Add category columns
        for cat in category_totals.keys():
            cat_lines = invoice_lines[invoice_lines['category_code'] == cat]
            row[cat] = cat_lines['line_total'].sum() if len(cat_lines) > 0 else 0.0

        invoice_rows.append(row)

    df_report = pd.DataFrame(invoice_rows)

    # Export to Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_file = os.path.join(OUTPUT_DIR, f"GFS_Corrected_{period_id}_{datetime.now().strftime('%Y%m%d')}.xlsx")

    print()
    print(f"💾 Exporting to: {output_file}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{period_id}"

    # Write headers
    headers = ['Invoice #', 'Date', 'Vendor'] + sorted(category_totals.keys()) + ['Total Invoice Amount']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_idx, row_data in enumerate(invoice_rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, 0.0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if header not in ['Invoice #', 'Date', 'Vendor']:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Add totals row
    total_row = len(invoice_rows) + 2
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)

    for col_idx, header in enumerate(headers, start=1):
        if header not in ['Invoice #', 'Date', 'Vendor']:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            formula = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
            cell = ws.cell(row=total_row, column=col_idx, value=formula)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_file)

    conn.close()

    print(f"✅ Report generated: {output_file}")
    print()
    print("=" * 80)

    return output_file

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python3 generate_corrected_gfs_reports.py <fiscal_period>")
        print("Example: python3 generate_corrected_gfs_reports.py FY26-P01")
        sys.exit(1)

    period = sys.argv[1]

    try:
        generate_report(period)
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

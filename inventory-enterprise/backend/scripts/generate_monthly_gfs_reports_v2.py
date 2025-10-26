#!/usr/bin/env python3
"""
GFS Monthly Reports V2 - With Correct Category Breakdowns
Uses shadow tables with corrected line items data
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import sys
from datetime import datetime
import os

DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'enterprise_inventory.db')
OUTPUT_DIR = os.path.expanduser('~/Desktop/GFS_Fiscal_Reports')

def generate_report(period_id):
    """Generate comprehensive report for a fiscal period"""

    print("=" * 80)
    print(f"GFS FINANCIAL REPORT V2 - {period_id}")
    print("=" * 80)
    print(f"Database: {DB_PATH}")
    print()

    # Connect to database
    conn = sqlite3.connect(DB_PATH)

    # Query invoice headers from shadow
    df_headers = pd.read_sql_query(f"""
        SELECT
            invoice_number,
            invoice_date,
            vendor,
            subtotal_cents / 100.0 as subtotal,
            gst_cents / 100.0 as gst,
            qst_cents / 100.0 as qst,
            total_cents / 100.0 as total,
            validation_status
        FROM invoice_headers_shadow
        WHERE fiscal_period_id = ?
        ORDER BY invoice_date, invoice_number
    """, conn, params=(period_id,))

    if len(df_headers) == 0:
        print(f"❌ No invoices found for period {period_id} in shadow tables")
        print(f"   Run reimport first: ./scripts/reimport_gfs_invoices_v2.sh --shadow --period {period_id}")
        conn.close()
        return None

    print(f"✓ Found {len(df_headers)} invoices")

    # Query line items with categories
    df_lines = pd.read_sql_query(f"""
        SELECT
            li.invoice_number,
            li.description,
            li.quantity_decimal as quantity,
            li.unit,
            li.unit_price_cents / 100.0 as unit_price,
            li.line_total_cents / 100.0 as line_total,
            li.category_code,
            ic.label as category_label,
            ic.gl_account,
            li.validation_status
        FROM invoice_line_items_shadow li
        LEFT JOIN item_categories ic ON li.category_code = ic.category_code
        WHERE li.invoice_number IN (
            SELECT invoice_number FROM invoice_headers_shadow WHERE fiscal_period_id = ?
        )
        ORDER BY li.invoice_number, li.description
    """, conn, params=(period_id,))

    print(f"✓ Found {len(df_lines)} line items")
    print()

    conn.close()

    # Category Summary
    category_summary = df_lines.groupby(['gl_account', 'category_label']).agg({
        'line_total': 'sum',
        'invoice_number': 'count'
    }).rename(columns={'invoice_number': 'line_count'}).round(2)

    category_summary = category_summary.reset_index()
    category_summary.columns = ['GL Account', 'Category', 'Total Amount', 'Line Count']

    # Invoice Summary
    invoice_summary = df_headers[['invoice_number', 'invoice_date', 'vendor', 'subtotal', 'gst', 'qst', 'total']]

    # Reconciliation Check
    line_sums = df_lines.groupby('invoice_number')['line_total'].sum()
    df_headers['line_items_sum'] = df_headers['invoice_number'].map(line_sums).fillna(0)
    df_headers['variance'] = (df_headers['subtotal'] - df_headers['line_items_sum']).abs()

    reconciliation = df_headers[['invoice_number', 'subtotal', 'line_items_sum', 'variance']]
    reconciliation.columns = ['Invoice Number', 'Header Subtotal', 'Line Items Sum', 'Variance']

    # Print summary
    print("SUMMARY STATISTICS")
    print("-" * 80)
    print(f"Total Invoices:        {len(df_headers)}")
    print(f"Total Line Items:      {len(df_lines)}")
    print(f"Total Amount:          ${df_headers['total'].sum():,.2f}")
    print(f"Total Subtotal:        ${df_headers['subtotal'].sum():,.2f}")
    print(f"Total GST:             ${df_headers['gst'].sum():,.2f}")
    print(f"Total QST:             ${df_headers['qst'].sum():,.2f}")
    print()

    print("CATEGORY BREAKDOWN")
    print("-" * 80)
    print(category_summary.to_string(index=False))
    print()

    # Validation issues
    issues = reconciliation[reconciliation['Variance'] > 0.50]
    if len(issues) > 0:
        print(f"⚠️  {len(issues)} invoices have variance >$0.50:")
        print(issues.to_string(index=False))
        print()

    # Export to Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_filename = os.path.join(OUTPUT_DIR, f"GFS_Report_V2_{period_id}_{datetime.now().strftime('%Y%m%d')}.xlsx")

    print(f"💾 Exporting to Excel: {output_filename}")

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = f"GFS Financial Report - {period_id}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_summary['A4'] = "Total Invoices:"
    ws_summary['B4'] = len(df_headers)
    ws_summary['A5'] = "Total Amount:"
    ws_summary['B5'] = df_headers['total'].sum()
    ws_summary['B5'].number_format = '$#,##0.00'

    ws_summary['A7'] = "Category"
    ws_summary['B7'] = "GL Account"
    ws_summary['C7'] = "Amount"
    ws_summary['D7'] = "Line Count"

    for col in ['A7', 'B7', 'C7', 'D7']:
        ws_summary[col].font = Font(bold=True)
        ws_summary[col].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    row = 8
    for _, cat_row in category_summary.iterrows():
        ws_summary[f'A{row}'] = cat_row['Category']
        ws_summary[f'B{row}'] = cat_row['GL Account']
        ws_summary[f'C{row}'] = cat_row['Total Amount']
        ws_summary[f'C{row}'].number_format = '$#,##0.00'
        ws_summary[f'D{row}'] = int(cat_row['Line Count'])
        row += 1

    # Sheet 2: Invoices
    ws_invoices = wb.create_sheet("Invoices")
    for r_idx, row_data in enumerate(invoice_summary.to_dict('records'), start=1):
        if r_idx == 1:
            # Header
            for c_idx, col_name in enumerate(invoice_summary.columns, start=1):
                cell = ws_invoices.cell(row=1, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Data
        for c_idx, (col_name, value) in enumerate(row_data.items(), start=1):
            cell = ws_invoices.cell(row=r_idx + 1, column=c_idx, value=value)
            if col_name in ['subtotal', 'gst', 'qst', 'total']:
                cell.number_format = '$#,##0.00'

    # Sheet 3: Categories
    ws_categories = wb.create_sheet("Categories")
    for r_idx, row_data in enumerate(category_summary.to_dict('records'), start=1):
        if r_idx == 1:
            for c_idx, col_name in enumerate(category_summary.columns, start=1):
                cell = ws_categories.cell(row=1, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for c_idx, (col_name, value) in enumerate(row_data.items(), start=1):
            cell = ws_categories.cell(row=r_idx + 1, column=c_idx, value=value)
            if col_name == 'Total Amount':
                cell.number_format = '$#,##0.00'

    # Sheet 4: Reconciliation
    ws_recon = wb.create_sheet("Reconciliation")
    for r_idx, row_data in enumerate(reconciliation.to_dict('records'), start=1):
        if r_idx == 1:
            for c_idx, col_name in enumerate(reconciliation.columns, start=1):
                cell = ws_recon.cell(row=1, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for c_idx, (col_name, value) in enumerate(row_data.items(), start=1):
            cell = ws_recon.cell(row=r_idx + 1, column=c_idx, value=value)
            if col_name in ['Header Subtotal', 'Line Items Sum', 'Variance']:
                cell.number_format = '$#,##0.00'

    # Adjust column widths
    for ws in [ws_summary, ws_invoices, ws_categories, ws_recon]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_filename)

    print(f"✅ Report generated successfully!")
    print()
    print("=" * 80)

    return output_filename


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 generate_monthly_gfs_reports_v2.py <fiscal_period>")
        print("Example: python3 generate_monthly_gfs_reports_v2.py FY26-P01")
        sys.exit(1)

    period = sys.argv[1]

    try:
        output_file = generate_report(period)
        if output_file:
            print(f"📁 Report saved: {output_file}")
            sys.exit(0)
        else:
            sys.exit(1)
    except Exception as e:
        print(f"❌ Error generating report: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

#!/usr/bin/env python3
"""
GFS Monthly Accounting Report Generator
Generates individual reports for each month with all GFS invoices
Accessible at any time - creates one file per month
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import glob
from calendar import monthrange

# ============================================================================
# CONFIGURATION
# ============================================================================

DATABASE_PATH = "../backend/database.db"
OUTPUT_DIR = "/Users/davidmikulis/Desktop/GFS_Monthly_Reports"
PDF_BASE_DIR = "/Users/davidmikulis/Desktop/Invoices"

# Cost code mapping
ITEM_CATEGORY_MAP = {
    'BAKE': '60110010 BAKE', 'BAKERY': '60110010 BAKE', 'BREAD': '60110010 BAKE',
    'ROLL': '60110010 BAKE', 'BUN': '60110010 BAKE', 'PASTRY': '60110010 BAKE',
    'BEV': '60110020 BEV + ECO', 'BEVERAGE': '60110020 BEV + ECO', 'JUICE': '60110020 BEV + ECO',
    'COFFEE': '60110020 BEV + ECO', 'TEA': '60110020 BEV + ECO', 'DRINK': '60110020 BEV + ECO',
    'MILK': '60110030 MILK', 'DAIRY': '60110030 MILK', 'CHEESE': '60110030 MILK',
    'YOGURT': '60110030 MILK', 'CREAM': '60110030 MILK', 'BUTTER': '60110030 MILK',
    'GROC': '60110040 GROC + MISC', 'GROCERY': '60110040 GROC + MISC', 'SAUCE': '60110040 GROC + MISC',
    'SPICE': '60110040 GROC + MISC', 'OIL': '60110040 GROC + MISC', 'PASTA': '60110040 GROC + MISC',
    'MEAT': '60110060 MEAT', 'BEEF': '60110060 MEAT', 'PORK': '60110060 MEAT',
    'CHICKEN': '60110060 MEAT', 'BACON': '60110060 MEAT', 'FISH': '60110060 MEAT',
    'PROD': '60110070 PROD', 'PRODUCE': '60110070 PROD', 'FRUIT': '60110070 PROD',
    'VEGETABLE': '60110070 PROD', 'LETTUCE': '60110070 PROD', 'TOMATO': '60110070 PROD',
    'CLEAN': '60220001 CLEAN', 'CLEANING': '60220001 CLEAN', 'SOAP': '60220001 CLEAN',
    'PAPER': '60260010 PAPER', 'TOWEL': '60260010 PAPER', 'NAPKIN': '60260010 PAPER',
    'EQUIP': '60665001 Small Equip', 'EQUIPMENT': '60665001 Small Equip',
    'LINEN': '60240010 LINEN', 'APRON': '60240010 LINEN',
    'PROPANE': '62869010 PROPANE', 'GAS': '62869010 PROPANE',
}

# 25-column schema
COLUMNS = [
    "Week Ending",
    "Vendor",
    "Date",
    "Invoice #",
    "60110010 BAKE",
    "60110020 BEV + ECO",
    "60110030 MILK",
    "60110040 GROC + MISC",
    "60110060 MEAT",
    "60110070 PROD",
    "60220001 CLEAN",
    "60260010 PAPER",
    "60665001 Small Equip",
    "62421100 FREIGHT",
    "60240010 LINEN",
    "62869010 PROPANE",
    "Other Costs",
    "63107000 GST",
    "63107100 QST",
    "Total Invoice Amount",
    "Total Food & Freight Reimb.",
    "Total Reimb. Other",
    "📎 Document Link",
    "Notes",
    "🔗 Copy/Paste Link"
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def map_item_to_cost_code(item_name):
    """Map item name to cost code category"""
    if not item_name:
        return 'Other Costs'

    item_upper = str(item_name).upper()
    for keyword, cost_code in ITEM_CATEGORY_MAP.items():
        if keyword in item_upper:
            return cost_code

    return 'Other Costs'

def check_pdf_exists(invoice_num, year, month):
    """Check if PDF exists in the monthly directory"""
    if not invoice_num:
        return False

    invoice_clean = str(invoice_num).replace(' ', '').replace('#', '')
    pdf_dir = os.path.join(PDF_BASE_DIR, f"{year}-{month:02d}")

    if not os.path.exists(pdf_dir):
        return False

    pattern = os.path.join(pdf_dir, f"*{invoice_clean}*.pdf")
    matches = glob.glob(pattern)

    return len(matches) > 0

def create_pdf_hyperlink(invoice_num, year, month, pdf_exists=True):
    """Create Excel HYPERLINK formula for PDF document"""
    if not invoice_num:
        return "NO PDF FOUND"

    invoice_clean = str(invoice_num).replace(' ', '').replace('#', '').split('.')[0]

    if not pdf_exists:
        return f"⚠️ Missing: {invoice_clean}"

    # Use file:// link to PDF on Desktop
    pdf_path = f"Users/davidmikulis/Desktop/Invoices/{year}-{month:02d}/{invoice_clean}.pdf"
    # Excel requires file:/// (three slashes) for local files
    url = f"file:///{pdf_path}"
    formula = f'=HYPERLINK("{url}","📄 {invoice_clean}")'

    return formula

def safe_float(val):
    """Convert value to float, return 0.0 if invalid"""
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def get_week_ending_date(year, month):
    """Get the last day of the month as week ending date"""
    last_day = monthrange(year, month)[1]
    return f"{year}-{month:02d}-{last_day:02d}"

def get_available_months(conn):
    """Query database to find all months with invoices"""
    query = """
    SELECT DISTINCT
        strftime('%Y', invoice_date) as year,
        strftime('%m', invoice_date) as month
    FROM processed_invoices
    WHERE invoice_date IS NOT NULL
    ORDER BY year, month
    """

    df = pd.read_sql_query(query, conn)
    months = []

    for _, row in df.iterrows():
        year = int(row['year'])
        month = int(row['month'])
        months.append((year, month))

    return months

# ============================================================================
# REPORT GENERATION
# ============================================================================

def generate_monthly_report(year, month):
    """Generate report for a specific month"""

    month_name = datetime(year, month, 1).strftime('%B')
    week_ending = get_week_ending_date(year, month)
    sheet_name = f"Week_Ending_{week_ending}"

    # Date range for the month
    start_date = f"{year}-{month:02d}-01"
    last_day = monthrange(year, month)[1]
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    print(f"\n{'='*80}")
    print(f"Generating Report: {month_name} {year}")
    print(f"{'='*80}")
    print(f"Date Range: {start_date} to {end_date}")
    print(f"Week Ending: {week_ending}")

    # Connect to database
    conn = sqlite3.connect(DATABASE_PATH)

    # Query invoices for this month
    query_invoices = """
    SELECT
        invoice_id,
        invoice_number,
        supplier,
        invoice_date,
        total_amount,
        gst,
        qst,
        subtotal
    FROM processed_invoices
    WHERE invoice_date >= ? AND invoice_date <= ?
    ORDER BY supplier, invoice_date, invoice_number
    """

    df_invoices = pd.read_sql_query(query_invoices, conn, params=(start_date, end_date))
    print(f"  ✓ Found {len(df_invoices)} invoices")

    if len(df_invoices) == 0:
        print(f"  ⚠️  No invoices for {month_name} {year}")
        conn.close()
        return None

    # Query invoice items
    query_items = """
    SELECT
        invoice_id,
        item_name,
        total_price
    FROM invoice_items
    WHERE invoice_id IN ({})
    """.format(','.join(['?'] * len(df_invoices)))

    invoice_ids = df_invoices['invoice_id'].tolist()
    df_items = pd.read_sql_query(query_items, conn, params=invoice_ids)
    print(f"  ✓ Found {len(df_items)} line items")

    conn.close()

    # Process invoices
    output_rows = []
    missing_pdf_count = 0

    for idx, invoice_row in df_invoices.iterrows():
        invoice_id = invoice_row['invoice_id']
        invoice_number = invoice_row['invoice_number']
        supplier = invoice_row['supplier'] or 'GFS'
        invoice_date = invoice_row['invoice_date']
        total_amount = safe_float(invoice_row['total_amount'])
        gst = safe_float(invoice_row['gst'])
        qst = safe_float(invoice_row['qst'])

        # Check if PDF exists
        pdf_exists = check_pdf_exists(invoice_number, year, month)
        if not pdf_exists:
            missing_pdf_count += 1

        # Get line items
        items = df_items[df_items['invoice_id'] == invoice_id]

        # Initialize cost codes
        cost_codes = {col: 0.0 for col in COLUMNS if col not in [
            'Week Ending', 'Vendor', 'Date', 'Invoice #',
            'Total Invoice Amount', 'Total Food & Freight Reimb.',
            'Total Reimb. Other', '📎 Document Link', 'Notes', '🔗 Copy/Paste Link',
            '63107000 GST', '63107100 QST'
        ]}

        # Map line items to cost codes
        for _, item_row in items.iterrows():
            item_name = item_row['item_name']
            item_price = safe_float(item_row['total_price'])
            cost_code = map_item_to_cost_code(item_name)
            cost_codes[cost_code] += item_price

        # Set taxes
        cost_codes['63107000 GST'] = gst
        cost_codes['63107100 QST'] = qst

        # Calculate totals
        food_freight = sum([
            cost_codes.get('60110010 BAKE', 0),
            cost_codes.get('60110020 BEV + ECO', 0),
            cost_codes.get('60110030 MILK', 0),
            cost_codes.get('60110040 GROC + MISC', 0),
            cost_codes.get('60110060 MEAT', 0),
            cost_codes.get('60110070 PROD', 0),
            cost_codes.get('60220001 CLEAN', 0),
            cost_codes.get('60260010 PAPER', 0),
            cost_codes.get('60665001 Small Equip', 0),
            cost_codes.get('62421100 FREIGHT', 0),
        ])

        reimb_other = sum([
            cost_codes.get('60240010 LINEN', 0),
            cost_codes.get('62869010 PROPANE', 0),
            cost_codes.get('Other Costs', 0),
        ])

        # Create notes
        notes = "⚠️ Missing PDF" if not pdf_exists else ""

        # Create row
        row = {
            'Week Ending': week_ending,
            'Vendor': supplier,
            'Date': invoice_date,
            'Invoice #': invoice_number,
            **cost_codes,
            'Total Invoice Amount': round(total_amount, 2),
            'Total Food & Freight Reimb.': round(food_freight, 2),
            'Total Reimb. Other': round(reimb_other, 2),
            '📎 Document Link': create_pdf_hyperlink(invoice_number, year, month, pdf_exists),
            'Notes': notes,
            '🔗 Copy/Paste Link': ''
        }

        output_rows.append(row)

    print(f"  ✓ Processed {len(output_rows)} invoice rows")
    if missing_pdf_count > 0:
        print(f"  ⚠️  Missing PDFs: {missing_pdf_count}")

    # Create dataframe
    df_output = pd.DataFrame(output_rows, columns=COLUMNS)

    # Round numeric columns
    numeric_cols = [col for col in COLUMNS if col not in [
        'Week Ending', 'Vendor', 'Date', 'Invoice #', '📎 Document Link', 'Notes', '🔗 Copy/Paste Link'
    ]]
    for col in numeric_cols:
        df_output[col] = df_output[col].apply(lambda x: round(float(x), 2) if pd.notna(x) else 0.0)

    # Export to Excel
    output_filename = os.path.join(OUTPUT_DIR, f"GFS_Accounting_{year}_{month:02d}_{month_name}.xlsx")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write headers
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_idx, row_data in enumerate(df_output.to_dict('records'), start=2):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)

                if col_name == '🔗 Copy/Paste Link':
                    cell.value = "Select row and Ctrl+C to copy"
                    cell.font = Font(italic=True, size=9, color="666666")
                elif col_name == '📎 Document Link':
                    cell.value = value
                    if '⚠️ Missing' in str(value):
                        cell.font = Font(color="FF0000", bold=True)
                elif col_name in numeric_cols:
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.value = value

        # Add grand total row
        total_row_idx = len(df_output) + 2
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws.cell(row=total_row_idx, column=1, value="GRAND TOTAL").font = total_font
        ws.cell(row=total_row_idx, column=1).fill = total_fill

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            if col_name in numeric_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                formula = f'=SUM({col_letter}2:{col_letter}{total_row_idx-1})'
                cell.value = formula
                cell.font = total_font
                cell.fill = total_fill
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

        # Adjust column widths
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name == '📎 Document Link':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
            elif col_name == '🔗 Copy/Paste Link':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
            elif col_name == 'Notes':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            elif col_name in ['Vendor', 'Invoice #']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            elif col_name in numeric_cols:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

        # Save workbook
        wb.save(output_filename)
        print(f"  ✓ Saved: {output_filename}")

        # Return summary
        total_amount = df_output['Total Invoice Amount'].sum()
        return {
            'year': year,
            'month': month,
            'month_name': month_name,
            'invoice_count': len(df_output),
            'total_amount': total_amount,
            'missing_pdfs': missing_pdf_count,
            'filename': output_filename
        }

    except Exception as e:
        print(f"  ✗ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

# ============================================================================
# MAIN FUNCTION
# ============================================================================

def generate_all_monthly_reports():
    """Generate reports for all available months"""

    print("=" * 80)
    print("GFS MONTHLY ACCOUNTING REPORT GENERATOR")
    print("=" * 80)
    print()

    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Output Directory: {OUTPUT_DIR}")

    # Connect to database
    print("\nConnecting to database...")
    conn = sqlite3.connect(DATABASE_PATH)

    # Get all available months
    print("Scanning database for available months...")
    available_months = get_available_months(conn)
    conn.close()

    if len(available_months) == 0:
        print("✗ No invoices found in database")
        return

    print(f"✓ Found invoices in {len(available_months)} months:")
    for year, month in available_months:
        month_name = datetime(year, month, 1).strftime('%B')
        print(f"  • {month_name} {year}")

    # Generate reports
    print("\n" + "=" * 80)
    print("GENERATING MONTHLY REPORTS")
    print("=" * 80)

    summaries = []
    for year, month in available_months:
        summary = generate_monthly_report(year, month)
        if summary:
            summaries.append(summary)

    # Final summary
    print("\n" + "=" * 80)
    print("GENERATION COMPLETE")
    print("=" * 80)
    print()

    if summaries:
        total_invoices = sum(s['invoice_count'] for s in summaries)
        total_amount = sum(s['total_amount'] for s in summaries)
        total_missing = sum(s['missing_pdfs'] for s in summaries)

        print(f"✅ Generated {len(summaries)} monthly reports")
        print(f"📊 Total Invoices: {total_invoices}")
        print(f"💰 Total Amount: ${total_amount:,.2f}")
        if total_missing > 0:
            print(f"⚠️  Missing PDFs: {total_missing}")
        print()

        print("MONTHLY BREAKDOWN:")
        print("-" * 80)
        for s in summaries:
            print(f"  {s['month_name']} {s['year']:4d}  |  {s['invoice_count']:3d} invoices  |  ${s['total_amount']:12,.2f}")
        print("-" * 80)
        print()

        print(f"📁 All reports saved to: {OUTPUT_DIR}")
        print()
        print("USAGE:")
        print("  • Each month has its own Excel file")
        print("  • Open any month's file to view/copy invoices")
        print("  • Click PDF links to view source documents")
        print("  • Copy/paste rows directly into master workbook")
    else:
        print("✗ No reports generated")

if __name__ == "__main__":
    import sys
    try:
        generate_all_monthly_reports()
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

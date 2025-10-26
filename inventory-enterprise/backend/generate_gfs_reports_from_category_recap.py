#!/usr/bin/env python3
"""
GFS Reports from Category Recap - Extract categories from PDF text
Parses the "CATEGORY RECAP" section that GFS includes in their invoices
This gives us accurate category breakdowns without relying on corrupted line items
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import re

DB_PATH = "data/enterprise_inventory.db"
OUTPUT_DIR = os.path.expanduser("~/Desktop/GFS_Fiscal_Reports_WITH_CATEGORIES")

# Map GFS category names to our GL account codes
GFS_CATEGORY_TO_GL = {
    'Produce': '60110070 PROD',
    'Meat': '60110060 MEAT',
    'Poultry': '60110060 MEAT',
    'Seafood': '60110060 MEAT',
    'Dairy': '60110030 MILK',
    'Frozen': '60110040 GROC + MISC',
    'Grocery': '60110040 GROC + MISC',
    'Beverage': '60110020 BEV + ECO',
    'Disposables': '60260010 PAPER',
    'Chemical': '60220001 CLEAN',
    'Tabletop': '60665001 Small Equip',
    'Fuel Charge': '62421100 FREIGHT',
}

# Standard columns for report
STANDARD_CATEGORIES = [
    '60110010 BAKE',
    '60110020 BEV + ECO',
    '60110030 MILK',
    '60110040 GROC + MISC',
    '60110060 MEAT',
    '60110070 PROD',
    '60220001 CLEAN',
    '60260010 PAPER',
    '60665001 Small Equip',
    '62421100 FREIGHT',
    '60240010 LINEN',
    '62869010 PROPANE',
    'Other Costs',
    '63107000 GST',
    '63107100 QST',
]

def parse_category_recap(text):
    """
    Extract category totals from GFS invoice text

    Handles multiple formats:
    1. CATEGORY RECAP section (multi-line)
    2. Inline categories before totals (e.g., "Beverage22.492318.41...")

    Returns: dict of {category_name: amount}
    """
    if not text:
        return {}

    categories = {}

    # Strategy 1: Look for CATEGORY RECAP section
    recap_match = re.search(r'CATEGORY\s+RECAP(.*?)(?:Total|Sub total|Product Total|\n\n|$)', text, re.IGNORECASE | re.DOTALL)

    if recap_match:
        recap_text = recap_match.group(1)
        lines = recap_text.strip().split('\n')

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Pattern: word(s) followed by numbers with decimals
            match = re.match(r'^([A-Za-z\s]+?)\s+([\d,]+\.[\d]{2})', line)

            if match:
                category_name = match.group(1).strip()
                amount_str = match.group(2).replace(',', '')

                try:
                    amount = float(amount_str)
                    categories[category_name] = amount
                except ValueError:
                    continue

    # Strategy 2: Look for CATEGORY RECAP with inline numbers
    # Example: "Produce2824.072.0001137.18" (no spaces between category and numbers)
    if not categories:
        # Find CATEGORY RECAP section (may not have newlines after)
        recap_section = re.search(r'CATEGORY\s+RECAP.*?(?:Product Total|Sub total)', text, re.IGNORECASE | re.DOTALL)

        if recap_section:
            section_text = recap_section.group(0)

            # Known GFS category names
            category_names = [
                'Produce', 'Meat', 'Poultry', 'Seafood', 'Dairy',
                'Frozen', 'Grocery', 'Beverage', 'Disposables',
                'Tabletop', 'Chemical', 'Fuel Charge'
            ]

            for category_name in category_names:
                # Pattern: CategoryName followed immediately by number with decimal
                # Example: "Produce2824.07" or "Meat10432.18"
                pattern = rf'{category_name}([\d]+\.[\d]{{2}})'
                match = re.search(pattern, section_text, re.IGNORECASE)

                if match:
                    amount_str = match.group(1)
                    try:
                        amount = float(amount_str)
                        categories[category_name] = amount
                    except ValueError:
                        continue

    return categories

def parse_taxes_from_text(text):
    """Extract GST and QST from invoice text"""
    gst = 0.0
    qst = 0.0

    if not text:
        return gst, qst

    # Look for GST/HST
    gst_match = re.search(r'(?:GST\/HST|GST)\s+\$?([\d,]+\.[\d]{2})', text, re.IGNORECASE)
    if gst_match:
        try:
            gst = float(gst_match.group(1).replace(',', ''))
        except:
            pass

    # Look for PST/QST
    qst_match = re.search(r'(?:PST\/QST|QST|PST)\s+\$?([\d,]+\.[\d]{2})', text, re.IGNORECASE)
    if qst_match:
        try:
            qst = float(qst_match.group(1).replace(',', ''))
        except:
            pass

    return gst, qst

def generate_report(period_id):
    """Generate report with categories from CATEGORY RECAP"""

    print("=" * 80)
    print(f"GFS REPORT WITH CATEGORIES - {period_id}")
    print("=" * 80)
    print()

    conn = sqlite3.connect(DB_PATH)

    # Get invoices with extracted text
    df_invoices = pd.read_sql_query("""
        SELECT
            d.id as document_id,
            d.invoice_number,
            d.invoice_date,
            d.vendor,
            d.invoice_amount as total_amount,
            d.extracted_text
        FROM documents d
        WHERE d.fiscal_period_id = ?
          AND d.mime_type = 'application/pdf'
          AND d.deleted_at IS NULL
        ORDER BY d.invoice_date, d.invoice_number
    """, conn, params=(period_id,))

    conn.close()

    if len(df_invoices) == 0:
        print(f"❌ No invoices found for period {period_id}")
        return None

    print(f"✓ Found {len(df_invoices)} invoices")
    print()

    # Process each invoice
    invoice_rows = []
    successful_parses = 0
    failed_parses = 0
    total_category_sum = 0.0
    total_invoice_sum = 0.0

    for idx, inv in df_invoices.iterrows():
        document_id = inv['document_id']
        invoice_num = inv['invoice_number']
        total_amount = float(inv['total_amount'])
        text = inv['extracted_text']

        # Parse category recap
        categories = parse_category_recap(text)
        gst, qst = parse_taxes_from_text(text)

        # Map to GL codes
        row = {
            'Fiscal Period': period_id,
            'Week Ending': '',  # Will be filled later
            'Vendor': inv['vendor'] or 'GFS',
            'Date': inv['invoice_date'],
            'Invoice #': invoice_num,
        }

        # Initialize all category columns to 0
        for cat in STANDARD_CATEGORIES:
            row[cat] = 0.0

        # Fill in parsed categories
        category_total = 0.0
        for gfs_cat, amount in categories.items():
            gl_code = GFS_CATEGORY_TO_GL.get(gfs_cat, 'Other Costs')
            row[gl_code] = row.get(gl_code, 0.0) + amount
            category_total += amount

        # Add taxes
        row['63107000 GST'] = gst
        row['63107100 QST'] = qst

        # Calculate totals
        food_freight = sum([
            row['60110010 BAKE'],
            row['60110020 BEV + ECO'],
            row['60110030 MILK'],
            row['60110040 GROC + MISC'],
            row['60110060 MEAT'],
            row['60110070 PROD'],
            row['60220001 CLEAN'],
            row['60260010 PAPER'],
            row['60665001 Small Equip'],
            row['62421100 FREIGHT'],
        ])

        reimb_other = sum([
            row['60240010 LINEN'],
            row['62869010 PROPANE'],
            row['Other Costs'],
        ])

        row['Total Invoice Amount'] = total_amount
        row['Total Food & Freight Reimb.'] = food_freight
        row['Total Reimb. Other'] = reimb_other
        row['📎 Document Link'] = f'=HYPERLINK("http://localhost:8083/api/owner/pdfs/{document_id}/preview","📄 {invoice_num}")'
        row['Notes'] = ''

        # Check if parsing was successful
        if category_total > 0:
            successful_parses += 1
            total_category_sum += category_total

            # Calculate variance
            # Category total + taxes should approximately equal invoice total
            calculated_total = category_total + gst + qst
            variance = abs(calculated_total - total_amount)

            if variance > 1.0:  # More than $1 difference
                row['Notes'] = f'Variance: ${variance:.2f}'

            print(f"✓ {invoice_num}: ${total_amount:,.2f} - Parsed {len(categories)} categories (${category_total:,.2f})")
        else:
            failed_parses += 1
            # Put entire amount in Other Costs if no category recap found
            row['Other Costs'] = total_amount - gst - qst
            row['Notes'] = 'No category recap found in PDF text'
            print(f"⚠️  {invoice_num}: ${total_amount:,.2f} - No category recap found")

        total_invoice_sum += total_amount
        invoice_rows.append(row)

    print()
    print("PARSING SUMMARY:")
    print("-" * 80)
    print(f"Successfully parsed: {successful_parses} invoices")
    print(f"Failed to parse:     {failed_parses} invoices")
    print(f"Success rate:        {successful_parses / len(df_invoices) * 100:.1f}%")
    print()

    # Category summary
    print("CATEGORY BREAKDOWN:")
    print("-" * 80)

    category_totals = {}
    for row in invoice_rows:
        for cat in STANDARD_CATEGORIES:
            if cat in row:
                category_totals[cat] = category_totals.get(cat, 0.0) + row[cat]

    for cat, total in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
        if total > 0:
            pct = (total / total_invoice_sum) * 100
            print(f"  {cat:30s}  ${total:12,.2f}  ({pct:5.1f}%)")

    print("-" * 80)
    print(f"  {'TOTAL':30s}  ${sum(category_totals.values()):12,.2f}")
    print()
    print(f"Invoice Total:   ${total_invoice_sum:,.2f}")
    print(f"Category Total:  ${sum(category_totals.values()):12,.2f}")
    print(f"Variance:        ${abs(sum(category_totals.values()) - total_invoice_sum):,.2f}")
    print()

    # Export to Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_file = os.path.join(OUTPUT_DIR, f"GFS_WithCategories_{period_id}_{datetime.now().strftime('%Y%m%d')}.xlsx")

    print(f"💾 Exporting to: {output_file}")

    # Create Excel with proper formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = period_id[:31]

    # Define columns
    columns = ['Fiscal Period', 'Week Ending', 'Vendor', 'Date', 'Invoice #'] + STANDARD_CATEGORIES + ['Total Invoice Amount', 'Total Food & Freight Reimb.', 'Total Reimb. Other', '📎 Document Link', 'Notes']

    # Write headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_idx, row_data in enumerate(invoice_rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_name in STANDARD_CATEGORIES + ['Total Invoice Amount', 'Total Food & Freight Reimb.', 'Total Reimb. Other']:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

    # Add grand total row
    total_row = len(invoice_rows) + 2
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    for col_idx, col_name in enumerate(columns, start=1):
        if col_name in STANDARD_CATEGORIES + ['Total Invoice Amount', 'Total Food & Freight Reimb.', 'Total Reimb. Other']:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            formula = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
            cell = ws.cell(row=total_row, column=col_idx, value=formula)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col_idx, col_name in enumerate(columns, start=1):
        if col_name == '📎 Document Link':
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
        elif col_name == 'Notes':
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 40
        elif col_name in ['Vendor', 'Invoice #', 'Fiscal Period']:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        elif col_name in STANDARD_CATEGORIES + ['Total Invoice Amount', 'Total Food & Freight Reimb.', 'Total Reimb. Other']:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    wb.save(output_file)

    print(f"✅ Report generated: {output_file}")
    print()
    print("=" * 80)

    return output_file

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python3 generate_gfs_reports_from_category_recap.py <fiscal_period>")
        print("Example: python3 generate_gfs_reports_from_category_recap.py FY26-P01")
        print()
        print("Or generate all periods:")
        print("  python3 generate_gfs_reports_from_category_recap.py --all")
        sys.exit(1)

    if sys.argv[1] == '--all':
        # Generate for all periods
        conn = sqlite3.connect(DB_PATH)
        periods = pd.read_sql_query("""
            SELECT DISTINCT fiscal_period_id
            FROM documents
            WHERE fiscal_period_id IS NOT NULL
              AND mime_type = 'application/pdf'
              AND deleted_at IS NULL
            ORDER BY fiscal_period_id
        """, conn)
        conn.close()

        for period_id in periods['fiscal_period_id']:
            try:
                generate_report(period_id)
                print()
            except Exception as e:
                print(f"❌ Error generating report for {period_id}: {e}")
                import traceback
                traceback.print_exc()
    else:
        period = sys.argv[1]
        try:
            generate_report(period)
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

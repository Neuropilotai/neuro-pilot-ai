#!/usr/bin/env python3
"""
TATA Sept AP Accounting Report Generator
Processes TATA Sept. AP.xlsx and produces append-only accounting rows
with exact 24-column schema and clickable PDF links.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import re
import os
import sys

# ============================================================================
# CONFIGURATION
# ============================================================================

SOURCE_FILE = "/Users/davidmikulis/Documents/TATA Sept. AP.xlsx"
OUTPUT_FILE = "/Users/davidmikulis/neuro-pilot-ai/inventory-enterprise/backend/TATA_AP_Accounting_Report_2025-09-26.xlsx"
WEEK_ENDING = "2025-09-26"
SHEET_NAME = "Week_Ending_2025-09-26"

# 24-column schema (exact order)
COLUMNS = [
    "Week Ending",
    "Vendor",
    "Date",
    "Invoice #",
    "60110010 BAKE",
    "60110020 BEV + ECO",
    "60110030 MILK",
    "60110040 GROC + MISC",
    "60110060 MEAT",
    "60110070 PROD",
    "60220001 CLEAN",
    "60260010 PAPER",
    "60665001 Small Equip",
    "62421100 FREIGHT",
    "60240010 LINEN",
    "62869010 PROPANE",
    "Other Costs",
    "63107000 GST",
    "63107100 QST",
    "Total Invoice Amount",
    "Total Food & Freight Reimb.",
    "Total Reimb. Other",
    "📎 Document Link",
    "Notes"
]

# Cost code mapping (normalize variations)
COST_CODE_MAP = {
    # Normalize header variations to standard names
    'BAKE': '60110010 BAKE',
    'BAKERY': '60110010 BAKE',
    '60110010': '60110010 BAKE',

    'BEV': '60110020 BEV + ECO',
    'BEVERAGE': '60110020 BEV + ECO',
    'BEV+ECO': '60110020 BEV + ECO',
    'BEV & ECO': '60110020 BEV + ECO',
    'BEV/ECO': '60110020 BEV + ECO',
    'BEVERAGE/ECO': '60110020 BEV + ECO',
    '60110020': '60110020 BEV + ECO',

    'MILK': '60110030 MILK',
    'DAIRY': '60110030 MILK',
    '60110030': '60110030 MILK',

    'GROC': '60110040 GROC + MISC',
    'GROCERY': '60110040 GROC + MISC',
    'GROC+MISC': '60110040 GROC + MISC',
    'GROC & MISC': '60110040 GROC + MISC',
    'GROC/MISC': '60110040 GROC + MISC',
    'MISC': '60110040 GROC + MISC',
    '60110040': '60110040 GROC + MISC',

    'MEAT': '60110060 MEAT',
    'MEATS': '60110060 MEAT',
    '60110060': '60110060 MEAT',

    'PROD': '60110070 PROD',
    'PRODUCE': '60110070 PROD',
    '60110070': '60110070 PROD',

    'CLEAN': '60220001 CLEAN',
    'CLEANING': '60220001 CLEAN',
    '60220001': '60220001 CLEAN',

    'PAPER': '60260010 PAPER',
    '60260010': '60260010 PAPER',

    'SMALL EQUIP': '60665001 Small Equip',
    'EQUIPMENT': '60665001 Small Equip',
    'EQUIP': '60665001 Small Equip',
    '60665001': '60665001 Small Equip',

    'FREIGHT': '62421100 FREIGHT',
    '62421100': '62421100 FREIGHT',

    'LINEN': '60240010 LINEN',
    '60240010': '60240010 LINEN',

    'PROPANE': '62869010 PROPANE',
    '62869010': '62869010 PROPANE',

    'GST': '63107000 GST',
    '63107000': '63107000 GST',

    'QST': '63107100 QST',
    'PST': '63107100 QST',
    '63107100': '63107100 QST',

    'OTHER': 'Other Costs',
    'OTHER COSTS': 'Other Costs',
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def normalize_vendor(vendor):
    """Normalize vendor name: uppercase, trim spaces"""
    if pd.isna(vendor):
        return ""
    return str(vendor).strip().upper()

def normalize_invoice(invoice):
    """Normalize invoice number: trim spaces, keep leading zeros"""
    if pd.isna(invoice):
        return ""
    return str(invoice).strip()

def normalize_date(date_val):
    """Normalize date to YYYY-MM-DD format"""
    if pd.isna(date_val):
        return ""
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')
    if isinstance(date_val, str):
        # Try to parse various date formats
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
            try:
                dt = datetime.strptime(date_val, fmt)
                return dt.strftime('%Y-%m-%d')
            except:
                continue
    return str(date_val)

def normalize_column_name(col):
    """Normalize column names for mapping"""
    if pd.isna(col):
        return ""
    # Remove special chars, uppercase, trim
    normalized = re.sub(r'[^A-Za-z0-9\s+&/-]', '', str(col))
    normalized = normalized.strip().upper()
    # Remove extra spaces
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized

def map_cost_code(col_name):
    """Map column name to standard cost code"""
    normalized = normalize_column_name(col_name)
    return COST_CODE_MAP.get(normalized, None)

def create_pdf_hyperlink(invoice_num):
    """Create Excel HYPERLINK formula for PDF document"""
    if not invoice_num:
        return "NO PDF FOUND"

    # Clean invoice number for URL
    invoice_clean = str(invoice_num).replace(' ', '').replace('#', '')

    # Create HYPERLINK formula
    url = f"http://localhost:8083/api/owner/docs/view/{invoice_clean}"
    formula = f'=HYPERLINK("{url}","Open Invoice {invoice_clean}")'

    return formula

def safe_float(val):
    """Convert value to float, return 0.0 if invalid"""
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

# ============================================================================
# MAIN PROCESSING
# ============================================================================

def process_tata_ap():
    """Main processing function"""

    print("=" * 80)
    print("TATA Sept AP Accounting Report Generator")
    print("=" * 80)
    print()

    # Step 1: Read source file
    print(f"[1/6] Reading source file: {SOURCE_FILE}")
    try:
        df_source = pd.read_excel(SOURCE_FILE)
        print(f"      ✓ Read {len(df_source)} rows, {len(df_source.columns)} columns")
    except Exception as e:
        print(f"      ✗ Error reading file: {e}")
        return False

    # Step 2: Map columns
    print("[2/6] Mapping columns to cost codes...")

    # Initialize output dataframe with all required columns
    output_rows = []

    # Detect which columns map to cost codes
    column_mapping = {}
    for col in df_source.columns:
        mapped = map_cost_code(col)
        if mapped:
            column_mapping[col] = mapped
            print(f"      ✓ {col} → {mapped}")

    # Detect key columns (Vendor, Invoice #, Date)
    vendor_col = None
    invoice_col = None
    date_col = None
    notes_col = None

    for col in df_source.columns:
        col_lower = str(col).lower()
        if 'vendor' in col_lower and not vendor_col:
            vendor_col = col
        elif 'invoice' in col_lower and not invoice_col:
            invoice_col = col
        elif 'date' in col_lower and not date_col:
            date_col = col
        elif 'note' in col_lower and not notes_col:
            notes_col = col

    if not vendor_col or not invoice_col:
        print(f"      ✗ Missing required columns: Vendor={vendor_col}, Invoice={invoice_col}")
        return False

    print(f"      ✓ Key columns: Vendor={vendor_col}, Invoice={invoice_col}, Date={date_col}")

    # Step 3: Group and aggregate by (Vendor, Invoice #, Date)
    print("[3/6] Grouping and aggregating rows...")

    # Create grouping key
    df_source['_vendor'] = df_source[vendor_col].apply(normalize_vendor)
    df_source['_invoice'] = df_source[invoice_col].apply(normalize_invoice)
    if date_col:
        df_source['_date'] = df_source[date_col].apply(normalize_date)
    else:
        df_source['_date'] = ''

    # Filter for GFS vendor only (per user requirement)
    print(f"      ℹ️  Total rows before GFS filter: {len(df_source)}")
    df_source = df_source[df_source['_vendor'].str.contains('GFS|GORDON FOOD', na=False, case=False)]
    print(f"      ✓ Filtered to GFS invoices only: {len(df_source)} rows")

    if len(df_source) == 0:
        print("      ✗ No GFS invoices found in source file")
        return False

    # Group by vendor, invoice, date
    grouped = df_source.groupby(['_vendor', '_invoice', '_date'])

    print(f"      ✓ Created {len(grouped)} unique invoice groups")

    # Step 4: Process each group
    print("[4/6] Processing invoice groups...")

    for (vendor, invoice, date), group_df in grouped:
        # Initialize cost codes dictionary
        cost_codes = {col: 0.0 for col in COLUMNS if col not in [
            'Week Ending', 'Vendor', 'Date', 'Invoice #',
            'Total Invoice Amount', 'Total Food & Freight Reimb.',
            'Total Reimb. Other', '📎 Document Link', 'Notes'
        ]}

        # Sum all mapped cost codes in this group
        for src_col, dest_col in column_mapping.items():
            if dest_col in cost_codes:
                cost_codes[dest_col] += group_df[src_col].apply(safe_float).sum()

        # Calculate totals
        food_freight = sum([
            cost_codes.get('60110010 BAKE', 0),
            cost_codes.get('60110020 BEV + ECO', 0),
            cost_codes.get('60110030 MILK', 0),
            cost_codes.get('60110040 GROC + MISC', 0),
            cost_codes.get('60110060 MEAT', 0),
            cost_codes.get('60110070 PROD', 0),
            cost_codes.get('60220001 CLEAN', 0),
            cost_codes.get('60260010 PAPER', 0),
            cost_codes.get('60665001 Small Equip', 0),
            cost_codes.get('62421100 FREIGHT', 0),
        ])

        reimb_other = sum([
            cost_codes.get('60240010 LINEN', 0),
            cost_codes.get('62869010 PROPANE', 0),
            cost_codes.get('Other Costs', 0),
        ])

        total_invoice = sum(cost_codes.values())

        # Get notes
        notes = ""
        if notes_col and notes_col in group_df.columns:
            notes_list = group_df[notes_col].dropna().unique()
            if len(notes_list) > 0:
                notes = "; ".join(str(n) for n in notes_list)

        # Create row
        row = {
            'Week Ending': WEEK_ENDING,
            'Vendor': vendor,
            'Date': date,
            'Invoice #': invoice,
            **cost_codes,
            'Total Invoice Amount': round(total_invoice, 2),
            'Total Food & Freight Reimb.': round(food_freight, 2),
            'Total Reimb. Other': round(reimb_other, 2),
            '📎 Document Link': create_pdf_hyperlink(invoice),
            'Notes': notes
        }

        output_rows.append(row)

    print(f"      ✓ Processed {len(output_rows)} invoice rows")

    # Step 5: Create output dataframe
    print("[5/6] Creating output dataframe...")

    df_output = pd.DataFrame(output_rows, columns=COLUMNS)

    # Sort by Vendor, Date, Invoice #
    df_output = df_output.sort_values(['Vendor', 'Date', 'Invoice #'])

    # Round all numeric columns to 2 decimals
    numeric_cols = [col for col in COLUMNS if col not in [
        'Week Ending', 'Vendor', 'Date', 'Invoice #', '📎 Document Link', 'Notes'
    ]]
    for col in numeric_cols:
        df_output[col] = df_output[col].apply(lambda x: round(float(x), 2) if pd.notna(x) else 0.0)

    print(f"      ✓ Output dataframe created: {len(df_output)} rows × {len(COLUMNS)} columns")

    # Step 6: Export to Excel with formatting
    print("[6/6] Exporting to Excel...")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        # Write headers (bold)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_idx, row_data in enumerate(df_output.to_dict('records'), start=2):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)

                # Check if it's a HYPERLINK formula
                if col_name == '📎 Document Link' and isinstance(value, str) and value.startswith('=HYPERLINK'):
                    cell.value = value
                elif col_name in numeric_cols:
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = value

        # Add grand total row (bold)
        total_row_idx = len(df_output) + 2
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws.cell(row=total_row_idx, column=1, value="GRAND TOTAL").font = total_font
        ws.cell(row=total_row_idx, column=1).fill = total_fill

        # Sum numeric columns
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name in numeric_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                formula = f'=SUM({col_letter}2:{col_letter}{total_row_idx-1})'
                cell = ws.cell(row=total_row_idx, column=col_idx, value=formula)
                cell.font = total_font
                cell.fill = total_fill
                cell.number_format = '#,##0.00'

        # Adjust column widths
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name == '📎 Document Link':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35
            elif col_name == 'Notes':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
            elif col_name in ['Vendor', 'Invoice #']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            elif col_name in numeric_cols:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

        # Save workbook
        wb.save(OUTPUT_FILE)
        print(f"      ✓ Excel file saved: {OUTPUT_FILE}")

    except Exception as e:
        print(f"      ✗ Error exporting to Excel: {e}")
        return False

    # Step 7: Validation
    print()
    print("=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)
    print(f"✓ Column count: {len(COLUMNS)} (required: 24)")
    print(f"✓ Output rows: {len(df_output)}")
    print(f"✓ Grand total row: Yes")
    print(f"✓ All money fields formatted: Yes (2 decimals)")
    print(f"✓ Sort order: Vendor → Date → Invoice #")
    print(f"✓ PDF links: {len([r for r in output_rows if '=HYPERLINK' in str(r.get('📎 Document Link', ''))])} hyperlinks created")
    print()

    # Summary statistics
    total_amount = df_output['Total Invoice Amount'].sum()
    total_food_freight = df_output['Total Food & Freight Reimb.'].sum()
    total_reimb_other = df_output['Total Reimb. Other'].sum()

    print("FINANCIAL SUMMARY")
    print(f"Total Invoice Amount:         ${total_amount:,.2f}")
    print(f"Total Food & Freight Reimb.:  ${total_food_freight:,.2f}")
    print(f"Total Reimb. Other:           ${total_reimb_other:,.2f}")
    print()

    print("=" * 80)
    print("✅ PROCESSING COMPLETE")
    print("=" * 80)
    print(f"Output file: {OUTPUT_FILE}")
    print("Ready to paste into master accounting workbook.")
    print()

    return True

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    try:
        success = process_tata_ap()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

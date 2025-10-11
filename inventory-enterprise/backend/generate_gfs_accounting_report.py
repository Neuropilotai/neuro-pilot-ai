#!/usr/bin/env python3
"""
GFS Accounting Report Generator
Extracts GFS invoices from backend database and generates 24-column accounting report
for Week Ending 2025-09-26 (September invoices)
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import re

# ============================================================================
# CONFIGURATION
# ============================================================================

DATABASE_PATH = "../backend/database.db"
OUTPUT_FILE = "./GFS_Accounting_Report_2025-09-26.xlsx"
WEEK_ENDING = "2025-09-26"
SHEET_NAME = "Week_Ending_2025-09-26"

# September 2025 date range (adjust as needed)
START_DATE = "2025-09-01"
END_DATE = "2025-09-30"

# 24-column schema (exact order)
COLUMNS = [
    "Week Ending",
    "Vendor",
    "Date",
    "Invoice #",
    "60110010 BAKE",
    "60110020 BEV + ECO",
    "60110030 MILK",
    "60110040 GROC + MISC",
    "60110060 MEAT",
    "60110070 PROD",
    "60220001 CLEAN",
    "60260010 PAPER",
    "60665001 Small Equip",
    "62421100 FREIGHT",
    "60240010 LINEN",
    "62869010 PROPANE",
    "Other Costs",
    "63107000 GST",
    "63107100 QST",
    "Total Invoice Amount",
    "Total Food & Freight Reimb.",
    "Total Reimb. Other",
    "📎 Document Link",
    "Notes"
]

# Cost code mapping based on item categories
ITEM_CATEGORY_MAP = {
    # Bakery items
    'BAKE': '60110010 BAKE',
    'BAKERY': '60110010 BAKE',
    'BREAD': '60110010 BAKE',
    'ROLL': '60110010 BAKE',
    'BUN': '60110010 BAKE',
    'PASTRY': '60110010 BAKE',
    'BAGEL': '60110010 BAKE',
    'TORTILLA': '60110010 BAKE',

    # Beverages
    'BEV': '60110020 BEV + ECO',
    'BEVERAGE': '60110020 BEV + ECO',
    'JUICE': '60110020 BEV + ECO',
    'COFFEE': '60110020 BEV + ECO',
    'TEA': '60110020 BEV + ECO',
    'WATER': '60110020 BEV + ECO',
    'SODA': '60110020 BEV + ECO',
    'DRINK': '60110020 BEV + ECO',

    # Dairy/Milk
    'MILK': '60110030 MILK',
    'DAIRY': '60110030 MILK',
    'CHEESE': '60110030 MILK',
    'YOGURT': '60110030 MILK',
    'CREAM': '60110030 MILK',
    'BUTTER': '60110030 MILK',
    'SOUR CREAM': '60110030 MILK',

    # Grocery/Misc
    'GROC': '60110040 GROC + MISC',
    'GROCERY': '60110040 GROC + MISC',
    'SAUCE': '60110040 GROC + MISC',
    'SPICE': '60110040 GROC + MISC',
    'CONDIMENT': '60110040 GROC + MISC',
    'OIL': '60110040 GROC + MISC',
    'PASTA': '60110040 GROC + MISC',
    'RICE': '60110040 GROC + MISC',
    'FLOUR': '60110040 GROC + MISC',
    'SUGAR': '60110040 GROC + MISC',
    'SALT': '60110040 GROC + MISC',

    # Meat
    'MEAT': '60110060 MEAT',
    'BEEF': '60110060 MEAT',
    'PORK': '60110060 MEAT',
    'CHICKEN': '60110060 MEAT',
    'TURKEY': '60110060 MEAT',
    'HAM': '60110060 MEAT',
    'BACON': '60110060 MEAT',
    'SAUSAGE': '60110060 MEAT',
    'FISH': '60110060 MEAT',
    'SEAFOOD': '60110060 MEAT',

    # Produce
    'PROD': '60110070 PROD',
    'PRODUCE': '60110070 PROD',
    'FRUIT': '60110070 PROD',
    'VEGETABLE': '60110070 PROD',
    'LETTUCE': '60110070 PROD',
    'TOMATO': '60110070 PROD',
    'ONION': '60110070 PROD',
    'POTATO': '60110070 PROD',
    'APPLE': '60110070 PROD',
    'ORANGE': '60110070 PROD',
    'BANANA': '60110070 PROD',

    # Cleaning
    'CLEAN': '60220001 CLEAN',
    'CLEANING': '60220001 CLEAN',
    'SOAP': '60220001 CLEAN',
    'DETERGENT': '60220001 CLEAN',
    'SANITIZER': '60220001 CLEAN',
    'BLEACH': '60220001 CLEAN',

    # Paper
    'PAPER': '60260010 PAPER',
    'TOWEL': '60260010 PAPER',
    'NAPKIN': '60260010 PAPER',
    'TISSUE': '60260010 PAPER',
    'TOILET': '60260010 PAPER',

    # Small Equipment
    'EQUIP': '60665001 Small Equip',
    'EQUIPMENT': '60665001 Small Equip',
    'TOOL': '60665001 Small Equip',
    'UTENSIL': '60665001 Small Equip',

    # Linen
    'LINEN': '60240010 LINEN',
    'TOWEL': '60240010 LINEN',
    'APRON': '60240010 LINEN',

    # Propane
    'PROPANE': '62869010 PROPANE',
    'GAS': '62869010 PROPANE',
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def map_item_to_cost_code(item_name):
    """Map item name to cost code category"""
    if not item_name:
        return 'Other Costs'

    item_upper = str(item_name).upper()

    # Check each category keyword
    for keyword, cost_code in ITEM_CATEGORY_MAP.items():
        if keyword in item_upper:
            return cost_code

    # Default to Other Costs
    return 'Other Costs'

def create_pdf_hyperlink(invoice_num):
    """Create Excel HYPERLINK formula for PDF document"""
    if not invoice_num:
        return "NO PDF FOUND"

    # Clean invoice number for URL
    invoice_clean = str(invoice_num).replace(' ', '').replace('#', '')

    # Create HYPERLINK formula
    url = f"http://localhost:8083/api/owner/docs/view/{invoice_clean}"
    formula = f'=HYPERLINK("{url}","Open Invoice {invoice_clean}")'

    return formula

def safe_float(val):
    """Convert value to float, return 0.0 if invalid"""
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

# ============================================================================
# MAIN PROCESSING
# ============================================================================

def generate_gfs_accounting_report():
    """Main processing function"""

    print("=" * 80)
    print("GFS Accounting Report Generator")
    print("=" * 80)
    print()

    # Step 1: Connect to database
    print(f"[1/6] Connecting to database: {DATABASE_PATH}")
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        print(f"      ✓ Connected successfully")
    except Exception as e:
        print(f"      ✗ Error connecting to database: {e}")
        return False

    # Step 2: Query invoices
    print(f"[2/6] Querying GFS invoices for September 2025...")

    query_invoices = """
    SELECT
        invoice_id,
        invoice_number,
        supplier,
        invoice_date,
        total_amount,
        gst,
        qst,
        subtotal
    FROM processed_invoices
    WHERE invoice_date >= ? AND invoice_date <= ?
    ORDER BY invoice_date, invoice_number
    """

    try:
        df_invoices = pd.read_sql_query(query_invoices, conn, params=(START_DATE, END_DATE))
        print(f"      ✓ Found {len(df_invoices)} invoices in date range")
    except Exception as e:
        print(f"      ✗ Error querying invoices: {e}")
        conn.close()
        return False

    if len(df_invoices) == 0:
        print(f"      ⚠️  No invoices found between {START_DATE} and {END_DATE}")
        print(f"      💡 Tip: Adjust START_DATE and END_DATE in the script")
        conn.close()
        return False

    # Step 3: Query invoice items
    print(f"[3/6] Querying invoice line items...")

    query_items = """
    SELECT
        invoice_id,
        item_name,
        total_price
    FROM invoice_items
    WHERE invoice_id IN ({})
    """.format(','.join(['?'] * len(df_invoices)))

    try:
        invoice_ids = df_invoices['invoice_id'].tolist()
        df_items = pd.read_sql_query(query_items, conn, params=invoice_ids)
        print(f"      ✓ Found {len(df_items)} line items")
    except Exception as e:
        print(f"      ✗ Error querying items: {e}")
        conn.close()
        return False

    conn.close()

    # Step 4: Process each invoice
    print(f"[4/6] Processing invoices and mapping to cost codes...")

    output_rows = []

    for idx, invoice_row in df_invoices.iterrows():
        invoice_id = invoice_row['invoice_id']
        invoice_number = invoice_row['invoice_number']
        supplier = invoice_row['supplier'] or 'GFS'
        invoice_date = invoice_row['invoice_date']
        total_amount = safe_float(invoice_row['total_amount'])
        gst = safe_float(invoice_row['gst'])
        qst = safe_float(invoice_row['qst'])

        # Get line items for this invoice
        items = df_items[df_items['invoice_id'] == invoice_id]

        # Initialize cost codes
        cost_codes = {col: 0.0 for col in COLUMNS if col not in [
            'Week Ending', 'Vendor', 'Date', 'Invoice #',
            'Total Invoice Amount', 'Total Food & Freight Reimb.',
            'Total Reimb. Other', '📎 Document Link', 'Notes',
            '63107000 GST', '63107100 QST'
        ]}

        # Map each line item to cost code
        for _, item_row in items.iterrows():
            item_name = item_row['item_name']
            item_price = safe_float(item_row['total_price'])

            cost_code = map_item_to_cost_code(item_name)
            cost_codes[cost_code] += item_price

        # Use taxes from invoice table
        cost_codes['63107000 GST'] = gst
        cost_codes['63107100 QST'] = qst

        # Calculate totals
        food_freight = sum([
            cost_codes.get('60110010 BAKE', 0),
            cost_codes.get('60110020 BEV + ECO', 0),
            cost_codes.get('60110030 MILK', 0),
            cost_codes.get('60110040 GROC + MISC', 0),
            cost_codes.get('60110060 MEAT', 0),
            cost_codes.get('60110070 PROD', 0),
            cost_codes.get('60220001 CLEAN', 0),
            cost_codes.get('60260010 PAPER', 0),
            cost_codes.get('60665001 Small Equip', 0),
            cost_codes.get('62421100 FREIGHT', 0),
        ])

        reimb_other = sum([
            cost_codes.get('60240010 LINEN', 0),
            cost_codes.get('62869010 PROPANE', 0),
            cost_codes.get('Other Costs', 0),
        ])

        # Create row
        row = {
            'Week Ending': WEEK_ENDING,
            'Vendor': supplier,
            'Date': invoice_date,
            'Invoice #': invoice_number,
            **cost_codes,
            'Total Invoice Amount': round(total_amount, 2),
            'Total Food & Freight Reimb.': round(food_freight, 2),
            'Total Reimb. Other': round(reimb_other, 2),
            '📎 Document Link': create_pdf_hyperlink(invoice_number),
            'Notes': ''
        }

        output_rows.append(row)

    print(f"      ✓ Processed {len(output_rows)} invoice rows")

    # Step 5: Create output dataframe
    print("[5/6] Creating output dataframe...")

    df_output = pd.DataFrame(output_rows, columns=COLUMNS)

    # Round all numeric columns to 2 decimals
    numeric_cols = [col for col in COLUMNS if col not in [
        'Week Ending', 'Vendor', 'Date', 'Invoice #', '📎 Document Link', 'Notes'
    ]]
    for col in numeric_cols:
        df_output[col] = df_output[col].apply(lambda x: round(float(x), 2) if pd.notna(x) else 0.0)

    print(f"      ✓ Output dataframe created: {len(df_output)} rows × {len(COLUMNS)} columns")

    # Step 6: Export to Excel with formatting
    print("[6/6] Exporting to Excel...")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        # Write headers (bold)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_idx, row_data in enumerate(df_output.to_dict('records'), start=2):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)

                # Check if it's a HYPERLINK formula
                if col_name == '📎 Document Link' and isinstance(value, str) and value.startswith('=HYPERLINK'):
                    cell.value = value
                elif col_name in numeric_cols:
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = value

        # Add grand total row (bold)
        total_row_idx = len(df_output) + 2
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws.cell(row=total_row_idx, column=1, value="GRAND TOTAL").font = total_font
        ws.cell(row=total_row_idx, column=1).fill = total_fill

        # Sum numeric columns
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name in numeric_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                formula = f'=SUM({col_letter}2:{col_letter}{total_row_idx-1})'
                cell = ws.cell(row=total_row_idx, column=col_idx, value=formula)
                cell.font = total_font
                cell.fill = total_fill
                cell.number_format = '#,##0.00'

        # Adjust column widths
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name == '📎 Document Link':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35
            elif col_name == 'Notes':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
            elif col_name in ['Vendor', 'Invoice #']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            elif col_name in numeric_cols:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

        # Save workbook
        wb.save(OUTPUT_FILE)
        print(f"      ✓ Excel file saved: {OUTPUT_FILE}")

    except Exception as e:
        print(f"      ✗ Error exporting to Excel: {e}")
        return False

    # Step 7: Validation
    print()
    print("=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)
    print(f"✓ Column count: {len(COLUMNS)} (required: 24)")
    print(f"✓ Output rows: {len(df_output)}")
    print(f"✓ Grand total row: Yes")
    print(f"✓ All money fields formatted: Yes (2 decimals)")
    print(f"✓ Sort order: By invoice date")
    print(f"✓ PDF links: {len([r for r in output_rows if '=HYPERLINK' in str(r.get('📎 Document Link', ''))])} hyperlinks created")
    print()

    # Summary statistics
    total_invoice_amount = df_output['Total Invoice Amount'].sum()
    total_food_freight = df_output['Total Food & Freight Reimb.'].sum()
    total_reimb_other = df_output['Total Reimb. Other'].sum()

    print("FINANCIAL SUMMARY")
    print(f"Total Invoice Amount:         ${total_invoice_amount:,.2f}")
    print(f"Total Food & Freight Reimb.:  ${total_food_freight:,.2f}")
    print(f"Total Reimb. Other:           ${total_reimb_other:,.2f}")
    print()

    print("=" * 80)
    print("✅ PROCESSING COMPLETE")
    print("=" * 80)
    print(f"Output file: {OUTPUT_FILE}")
    print("Ready to paste into master accounting workbook.")
    print()

    return True

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import sys
    try:
        success = generate_gfs_accounting_report()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

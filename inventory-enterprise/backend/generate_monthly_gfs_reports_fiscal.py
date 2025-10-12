#!/usr/bin/env python3
"""
GFS Monthly Accounting Report Generator - Fiscal Year Edition
Generates reports by fiscal period using FY25/FY26 calendar structure
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import glob

# ============================================================================
# CONFIGURATION
# ============================================================================

DATABASE_PATH = "data/enterprise_inventory.db"
OUTPUT_DIR = "/Users/davidmikulis/Desktop/GFS_Fiscal_Reports"
PDF_BASE_DIR = "/Users/davidmikulis/Desktop/Invoices"

# Cost code mapping
ITEM_CATEGORY_MAP = {
    'BAKE': '60110010 BAKE', 'BAKERY': '60110010 BAKE', 'BREAD': '60110010 BAKE',
    'ROLL': '60110010 BAKE', 'BUN': '60110010 BAKE', 'PASTRY': '60110010 BAKE',
    'BEV': '60110020 BEV + ECO', 'BEVERAGE': '60110020 BEV + ECO', 'JUICE': '60110020 BEV + ECO',
    'COFFEE': '60110020 BEV + ECO', 'TEA': '60110020 BEV + ECO', 'DRINK': '60110020 BEV + ECO',
    'MILK': '60110030 MILK', 'DAIRY': '60110030 MILK', 'CHEESE': '60110030 MILK',
    'YOGURT': '60110030 MILK', 'CREAM': '60110030 MILK', 'BUTTER': '60110030 MILK',
    'GROC': '60110040 GROC + MISC', 'GROCERY': '60110040 GROC + MISC', 'SAUCE': '60110040 GROC + MISC',
    'SPICE': '60110040 GROC + MISC', 'OIL': '60110040 GROC + MISC', 'PASTA': '60110040 GROC + MISC',
    'MEAT': '60110060 MEAT', 'BEEF': '60110060 MEAT', 'PORK': '60110060 MEAT',
    'CHICKEN': '60110060 MEAT', 'BACON': '60110060 MEAT', 'FISH': '60110060 MEAT',
    'PROD': '60110070 PROD', 'PRODUCE': '60110070 PROD', 'FRUIT': '60110070 PROD',
    'VEGETABLE': '60110070 PROD', 'LETTUCE': '60110070 PROD', 'TOMATO': '60110070 PROD',
    'CLEAN': '60220001 CLEAN', 'CLEANING': '60220001 CLEAN', 'SOAP': '60220001 CLEAN',
    'PAPER': '60260010 PAPER', 'TOWEL': '60260010 PAPER', 'NAPKIN': '60260010 PAPER',
    'EQUIP': '60665001 Small Equip', 'EQUIPMENT': '60665001 Small Equip',
    'LINEN': '60240010 LINEN', 'APRON': '60240010 LINEN',
    'PROPANE': '62869010 PROPANE', 'GAS': '62869010 PROPANE',
}

# 26-column schema (added Fiscal Period column)
COLUMNS = [
    "Fiscal Period",
    "Week Ending",
    "Vendor",
    "Date",
    "Invoice #",
    "60110010 BAKE",
    "60110020 BEV + ECO",
    "60110030 MILK",
    "60110040 GROC + MISC",
    "60110060 MEAT",
    "60110070 PROD",
    "60220001 CLEAN",
    "60260010 PAPER",
    "60665001 Small Equip",
    "62421100 FREIGHT",
    "60240010 LINEN",
    "62869010 PROPANE",
    "Other Costs",
    "63107000 GST",
    "63107100 QST",
    "Total Invoice Amount",
    "Total Food & Freight Reimb.",
    "Total Reimb. Other",
    "📎 Document Link",
    "Notes"
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def map_item_to_cost_code(item_name):
    """Map item name to cost code category"""
    if not item_name:
        return 'Other Costs'

    item_upper = str(item_name).upper()
    for keyword, cost_code in ITEM_CATEGORY_MAP.items():
        if keyword in item_upper:
            return cost_code

    return 'Other Costs'

def create_pdf_hyperlink(invoice_num):
    """Create Excel HYPERLINK formula for PDF document"""
    if not invoice_num:
        return "NO PDF FOUND"

    invoice_clean = str(invoice_num).replace(' ', '').replace('#', '').split('.')[0]

    # Use API endpoint for PDF viewing
    url = f"http://localhost:8083/api/owner/docs/view/{invoice_clean}"
    formula = f'=HYPERLINK("{url}","📄 {invoice_clean}")'

    return formula

def safe_float(val):
    """Convert value to float, return 0.0 if invalid"""
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def get_fiscal_periods(conn):
    """Query database to find all fiscal periods with invoices"""
    query = """
    SELECT DISTINCT
        fp.fiscal_year,
        fp.period,
        fp.fiscal_year_id,
        fp.period_start_date,
        fp.period_end_date,
        fp.notes as period_name
    FROM fiscal_periods fp
    JOIN documents d ON d.fiscal_year_id = fp.fiscal_year_id
        AND d.fiscal_period_id = 'FY' || (fp.fiscal_year % 100) || '-P' || printf('%02d', fp.period)
    WHERE d.mime_type = 'application/pdf'
      AND d.deleted_at IS NULL
    ORDER BY fp.fiscal_year, fp.period
    """

    df = pd.read_sql_query(query, conn)
    periods = []

    for _, row in df.iterrows():
        periods.append({
            'fiscal_year': int(row['fiscal_year']),
            'period': int(row['period']),
            'fiscal_year_id': row['fiscal_year_id'],
            'period_id': f"FY{int(row['fiscal_year']) % 100}-P{int(row['period']):02d}",
            'period_name': row['period_name'],
            'start_date': row['period_start_date'],
            'end_date': row['period_end_date']
        })

    return periods

# ============================================================================
# REPORT GENERATION
# ============================================================================

def generate_fiscal_period_report(period_info):
    """Generate report for a specific fiscal period"""

    fiscal_year = period_info['fiscal_year']
    period = period_info['period']
    period_id = period_info['period_id']
    period_name = period_info['period_name']
    start_date = period_info['start_date']
    end_date = period_info['end_date']

    sheet_name = f"{period_id}_{period_name.replace(' ', '_')}"

    print(f"\n{'='*80}")
    print(f"Generating Report: {period_id} - {period_name}")
    print(f"{'='*80}")
    print(f"Date Range: {start_date} to {end_date}")

    # Connect to database
    conn = sqlite3.connect(DATABASE_PATH)

    # Query invoices for this fiscal period
    query_invoices = """
    SELECT
        d.id as invoice_id,
        d.invoice_number,
        d.vendor as supplier,
        d.invoice_date,
        d.invoice_amount as total_amount,
        d.fiscal_year_id,
        d.fiscal_period_id
    FROM documents d
    WHERE d.fiscal_period_id = ?
      AND d.mime_type = 'application/pdf'
      AND d.deleted_at IS NULL
    ORDER BY d.invoice_date, d.invoice_number
    """

    df_invoices = pd.read_sql_query(query_invoices, conn, params=(period_id,))
    print(f"  ✓ Found {len(df_invoices)} invoices")

    if len(df_invoices) == 0:
        print(f"  ⚠️  No invoices for {period_id} - {period_name}")
        conn.close()
        return None

    # Query invoice line items from invoice_line_items table
    query_items = """
    SELECT
        ili.invoice_number,
        ili.description as item_name,
        ili.unit_price * ili.quantity as total_price
    FROM invoice_line_items ili
    WHERE ili.invoice_number IN ({})
    """.format(','.join(['?'] * len(df_invoices)))

    invoice_numbers = df_invoices['invoice_number'].tolist()
    df_items = pd.read_sql_query(query_items, conn, params=invoice_numbers)
    print(f"  ✓ Found {len(df_items)} line items")

    conn.close()

    # Process invoices
    output_rows = []

    for idx, invoice_row in df_invoices.iterrows():
        invoice_id = invoice_row['invoice_id']
        invoice_number = invoice_row['invoice_number']
        supplier = invoice_row['supplier'] or 'GFS'
        invoice_date = invoice_row['invoice_date']
        total_amount = safe_float(invoice_row['total_amount'])

        # Get line items
        items = df_items[df_items['invoice_number'] == invoice_number]

        # Initialize cost codes
        cost_codes = {col: 0.0 for col in COLUMNS if col not in [
            'Fiscal Period', 'Week Ending', 'Vendor', 'Date', 'Invoice #',
            'Total Invoice Amount', 'Total Food & Freight Reimb.',
            'Total Reimb. Other', '📎 Document Link', 'Notes',
            '63107000 GST', '63107100 QST'
        ]}

        # Map line items to cost codes
        for _, item_row in items.iterrows():
            item_name = item_row['item_name']
            item_price = safe_float(item_row['total_price'])
            cost_code = map_item_to_cost_code(item_name)
            cost_codes[cost_code] += item_price

        # Set taxes (0.0 for now - can be calculated from line items if needed)
        cost_codes['63107000 GST'] = 0.0
        cost_codes['63107100 QST'] = 0.0

        # Calculate totals
        food_freight = sum([
            cost_codes.get('60110010 BAKE', 0),
            cost_codes.get('60110020 BEV + ECO', 0),
            cost_codes.get('60110030 MILK', 0),
            cost_codes.get('60110040 GROC + MISC', 0),
            cost_codes.get('60110060 MEAT', 0),
            cost_codes.get('60110070 PROD', 0),
            cost_codes.get('60220001 CLEAN', 0),
            cost_codes.get('60260010 PAPER', 0),
            cost_codes.get('60665001 Small Equip', 0),
            cost_codes.get('62421100 FREIGHT', 0),
        ])

        reimb_other = sum([
            cost_codes.get('60240010 LINEN', 0),
            cost_codes.get('62869010 PROPANE', 0),
            cost_codes.get('Other Costs', 0),
        ])

        # Create row
        row = {
            'Fiscal Period': period_id,
            'Week Ending': end_date,  # Use fiscal period end date
            'Vendor': supplier,
            'Date': invoice_date,
            'Invoice #': invoice_number,
            **cost_codes,
            'Total Invoice Amount': round(total_amount, 2),
            'Total Food & Freight Reimb.': round(food_freight, 2),
            'Total Reimb. Other': round(reimb_other, 2),
            '📎 Document Link': create_pdf_hyperlink(invoice_number),
            'Notes': ''
        }

        output_rows.append(row)

    print(f"  ✓ Processed {len(output_rows)} invoice rows")

    # Create dataframe
    df_output = pd.DataFrame(output_rows, columns=COLUMNS)

    # Round numeric columns
    numeric_cols = [col for col in COLUMNS if col not in [
        'Fiscal Period', 'Week Ending', 'Vendor', 'Date', 'Invoice #', '📎 Document Link', 'Notes'
    ]]
    for col in numeric_cols:
        df_output[col] = df_output[col].apply(lambda x: round(float(x), 2) if pd.notna(x) else 0.0)

    # Export to Excel
    output_filename = os.path.join(OUTPUT_DIR, f"GFS_Accounting_{period_id}_{period_name.replace(' ', '_')}.xlsx")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        # Write headers
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_idx, row_data in enumerate(df_output.to_dict('records'), start=2):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)

                if col_name == '📎 Document Link':
                    cell.value = value
                elif col_name in numeric_cols:
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.value = value

        # Add grand total row
        total_row_idx = len(df_output) + 2
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws.cell(row=total_row_idx, column=1, value="GRAND TOTAL").font = total_font
        ws.cell(row=total_row_idx, column=1).fill = total_fill

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            if col_name in numeric_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                formula = f'=SUM({col_letter}2:{col_letter}{total_row_idx-1})'
                cell.value = formula
                cell.font = total_font
                cell.fill = total_fill
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

        # Adjust column widths
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name == '📎 Document Link':
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
            elif col_name in ['Vendor', 'Invoice #', 'Fiscal Period']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            elif col_name in numeric_cols:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

        # Save workbook
        wb.save(output_filename)
        print(f"  ✓ Saved: {output_filename}")

        # Return summary
        total_amount = df_output['Total Invoice Amount'].sum()
        return {
            'fiscal_year': fiscal_year,
            'period': period,
            'period_id': period_id,
            'period_name': period_name,
            'invoice_count': len(df_output),
            'total_amount': total_amount,
            'filename': output_filename
        }

    except Exception as e:
        print(f"  ✗ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

# ============================================================================
# MAIN FUNCTION
# ============================================================================

def generate_all_fiscal_reports():
    """Generate reports for all fiscal periods with invoices"""

    print("=" * 80)
    print("GFS FISCAL PERIOD ACCOUNTING REPORT GENERATOR")
    print("=" * 80)
    print()

    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Output Directory: {OUTPUT_DIR}")

    # Connect to database
    print("\nConnecting to database...")
    conn = sqlite3.connect(DATABASE_PATH)

    # Get all fiscal periods with invoices
    print("Scanning database for fiscal periods...")
    fiscal_periods = get_fiscal_periods(conn)
    conn.close()

    if len(fiscal_periods) == 0:
        print("✗ No invoices found in database")
        return

    print(f"✓ Found invoices in {len(fiscal_periods)} fiscal periods:")
    for period in fiscal_periods:
        print(f"  • {period['period_id']}: {period['period_name']} ({period['start_date']} to {period['end_date']})")

    # Generate reports
    print("\n" + "=" * 80)
    print("GENERATING FISCAL PERIOD REPORTS")
    print("=" * 80)

    summaries = []
    for period in fiscal_periods:
        summary = generate_fiscal_period_report(period)
        if summary:
            summaries.append(summary)

    # Final summary
    print("\n" + "=" * 80)
    print("GENERATION COMPLETE")
    print("=" * 80)
    print()

    if summaries:
        total_invoices = sum(s['invoice_count'] for s in summaries)
        total_amount = sum(s['total_amount'] for s in summaries)

        print(f"✅ Generated {len(summaries)} fiscal period reports")
        print(f"📊 Total Invoices: {total_invoices}")
        print(f"💰 Total Amount: ${total_amount:,.2f}")
        print()

        print("FISCAL PERIOD BREAKDOWN:")
        print("-" * 80)
        for s in summaries:
            print(f"  {s['period_id']} {s['period_name']:20s}  |  {s['invoice_count']:3d} invoices  |  ${s['total_amount']:12,.2f}")
        print("-" * 80)
        print()

        print(f"📁 All reports saved to: {OUTPUT_DIR}")
        print()
        print("USAGE:")
        print("  • Each fiscal period has its own Excel file")
        print("  • Reports are organized by Sodexo FY25/FY26 calendar")
        print("  • Open any period's file to view/copy invoices")
        print("  • Click PDF links to view source documents")
    else:
        print("✗ No reports generated")

if __name__ == "__main__":
    import sys
    try:
        generate_all_fiscal_reports()
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
